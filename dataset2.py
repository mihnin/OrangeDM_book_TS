import pandas as pd
import numpy as np
import datetime as dt
from dateutil.relativedelta import relativedelta
import random
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

def is_holiday_or_weekend(date):
    """Проверяет, является ли дата выходным или праздником в РФ"""
    # Выходные дни (суббота и воскресенье)
    if date.weekday() >= 5:
        return True
    
    # Основные российские праздники
    holidays = [
        # Новогодние каникулы
        (1, 1), (2, 1), (3, 1), (4, 1), (5, 1), (6, 1), (7, 1), (8, 1),
        # Рождество
        (7, 1),
        # День защитника отечества
        (23, 2),
        # Международный женский день
        (8, 3),
        # Праздник Весны и Труда
        (1, 5),
        # День Победы
        (9, 5),
        # День России
        (12, 6),
        # День народного единства
        (4, 11)
    ]
    
    # Проверка на праздники
    if (date.day, date.month) in holidays:
        return True
    
    # Дополнительно: перенос праздников, если они выпадают на выходные
    # (упрощенная логика, не учитывает все правила переносов)
    for day, month in holidays:
        holiday_date = dt.date(date.year, month, day)
        if holiday_date.weekday() >= 5:  # выпадает на выходной
            next_workday = holiday_date + dt.timedelta(days=(7 - holiday_date.weekday() + 1))
            if date == next_workday:
                return True
    
    return False

def generate_working_days(start_date, end_date):
    """Генерирует список рабочих дней между указанными датами"""
    working_days = []
    current_date = start_date
    while current_date <= end_date:
        if not is_holiday_or_weekend(current_date):
            working_days.append(current_date)
        current_date += dt.timedelta(days=1)
    return working_days

def generate_months(start_date, end_date):
    """Генерирует список с первыми днями каждого месяца между указанными датами"""
    months = []
    current_date = dt.date(start_date.year, start_date.month, 1)  # Первое число месяца
    
    while current_date <= end_date:
        months.append(current_date)
        # Переходим к первому числу следующего месяца
        current_date = (current_date + relativedelta(months=1))
    
    return months

def generate_time_series_with_trend_and_seasonality(dates, base_value, trend_factor, seasonal_amplitude, noise_level):
    """
    Генерирует временной ряд с трендом и сезонностью для помесячных данных
    
    Параметры:
    - dates: список дат (первые числа месяцев)
    - base_value: начальное значение
    - trend_factor: коэффициент тренда (годовой рост)
    - seasonal_amplitude: амплитуда сезонных колебаний
    - noise_level: уровень случайного шума
    """
    values = []
    for i, date in enumerate(dates):
        # Добавляем тренд (линейный рост)
        years_passed = ((date.year - dates[0].year) * 12 + date.month - dates[0].month) / 12
        trend = base_value * (1 + trend_factor) ** years_passed
        
        # Добавляем сезонность (годовая)
        month_of_year = date.month
        seasonality = seasonal_amplitude * np.sin(2 * np.pi * month_of_year / 12)
        
        # Добавляем случайный шум
        noise = np.random.normal(0, noise_level * trend)
        
        # Итоговое значение
        value = max(0, trend + seasonality * trend + noise)
        values.append(round(value, 2))
    
    return values

def generate_dataset():
    # Определяем период - 10 лет
    start_date = dt.date(2015, 1, 1)
    end_date = dt.date(2024, 12, 31)
    
    # Генерируем список месяцев
    months = generate_months(start_date, end_date)
    
    # Создаем базовый DataFrame с датами
    df = pd.DataFrame({'Дата': months})
    
    # Генерируем имена контрагентов
    debtor_names = [
        "ООО 'ТехноПром'", "АО 'Меркурий'", "ООО 'СтройИнвест'", 
        "ЗАО 'ЭнергоСбыт'", "ООО 'АгроХолдинг'", "ИП Иванов А.А.", 
        "ООО 'МеталлГрупп'", "АО 'ТрансЛогистика'", "ООО 'МедТех'", 
        "ЗАО 'ИТ-Решения'"
    ]
    
    creditor_names = [
        "ООО 'ПоставкаПлюс'", "АО 'ТехноСервис'", "ООО 'СырьеТорг'", 
        "ПАО 'ЭнергоСеть'", "ООО 'ЛогистикаПро'", "АО 'СтальИмпорт'", 
        "ООО 'ХимПром'", "ЗАО 'СтройМатериалы'", "ООО 'ФинансГрупп'", 
        "АО 'ТехИмпорт'"
    ]
    
    # Генерируем данные для материалов
    materials_base = 5_000_000
    df['Материалы'] = generate_time_series_with_trend_and_seasonality(
        months, materials_base, 0.08, 0.15, 0.02
    )
    
    # Генерируем данные для готовой продукции
    finished_goods_base = 7_000_000
    df['Готовая продукция'] = generate_time_series_with_trend_and_seasonality(
        months, finished_goods_base, 0.1, 0.2, 0.03
    )
    
    # Генерируем данные для дебиторской задолженности по контрагентам
    accounts_receivable_base = 10_000_000
    debtor_weights = [random.uniform(0.05, 0.2) for _ in range(10)]
    debtor_weights = [w/sum(debtor_weights) for w in debtor_weights]
    
    total_receivables = generate_time_series_with_trend_and_seasonality(
        months, accounts_receivable_base, 0.12, 0.1, 0.04
    )
    
    for i, name in enumerate(debtor_names):
        column_name = f"ДЗ: {name}"
        df[column_name] = [round(value * debtor_weights[i] * (1 + 0.1 * np.sin(j/3)), 2) 
                           for j, value in enumerate(total_receivables)]
    
    # Генерируем данные для кредиторской задолженности по контрагентам
    accounts_payable_base = 8_000_000
    creditor_weights = [random.uniform(0.05, 0.2) for _ in range(10)]
    creditor_weights = [w/sum(creditor_weights) for w in creditor_weights]
    
    total_payables = generate_time_series_with_trend_and_seasonality(
        months, accounts_payable_base, 0.07, 0.12, 0.03
    )
    
    for i, name in enumerate(creditor_names):
        column_name = f"КЗ: {name}"
        df[column_name] = [round(value * creditor_weights[i] * (1 + 0.1 * np.sin(j/4)), 2) 
                           for j, value in enumerate(total_payables)]
    
    # Добавляем итоговые столбцы
    df['Дебиторская задолженность ИТОГО'] = df[[col for col in df.columns if col.startswith('ДЗ:')]].sum(axis=1)
    df['Кредиторская задолженность ИТОГО'] = df[[col for col in df.columns if col.startswith('КЗ:')]].sum(axis=1)
    
    # Добавляем расчёт Запасов и ЧОК
    df['Запасы'] = df['Материалы'] + df['Готовая продукция']
    df['ЧОК'] = df['Запасы'] + df['Дебиторская задолженность ИТОГО'] - df['Кредиторская задолженность ИТОГО']
    
    # Форматируем столбец даты
    df['Дата'] = pd.to_datetime(df['Дата'])
    
    return df

def export_to_excel(df, filename="financial_dataset_monthly.xlsx"):
    """Экспортирует помесячные данные в Excel с форматированием"""
    file_path = os.path.join(r"C:\OrangeDM_book_TS", filename)
    
    # Создаем писателя Excel
    with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
        # Первая закладка - оригинальные данные
        df.to_excel(writer, sheet_name='Данные ЧОК помесячно', index=False)
        
        # Получаем доступ к листу для форматирования
        workbook = writer.book
        worksheet = writer.sheets['Данные ЧОК помесячно']
        
        # Явно форматируем столбец с датами - только месяц и год
        date_col_idx = df.columns.get_loc('Дата') + 1  # +1 для Excel индексации
        for row_idx in range(2, len(df) + 2):  # начиная со второй строки (после заголовка)
            cell = worksheet.cell(row=row_idx, column=1)  # Столбец Дата (A)
            cell.number_format = 'mm.yyyy'
        
        # Форматируем заголовки
        header_font = Font(bold=True)
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Применяем форматирование к заголовкам
        for col_idx, column in enumerate(df.columns):
            cell = worksheet.cell(row=1, column=col_idx+1)
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Автоподбор ширины столбцов
        for col_idx, column in enumerate(df.columns):
            column_width = max(len(str(column)), 15)
            worksheet.column_dimensions[get_column_letter(col_idx+1)].width = column_width
        
        # Форматирование числовых значений
        for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, max_row=len(df)+1, min_col=2, max_col=len(df.columns))):
            for cell in row:
                cell.number_format = '#,##0.00₽'
        
        # Вторая закладка - данные в четырех столбцах (Дата, Статья, Контрагент, Сумма)
        # Создаем копию DataFrame для преобразования
        df_long = pd.DataFrame(columns=['Дата', 'Статья', 'Контрагент', 'Сумма'])
        
        # Для каждого числового столбца создаем отдельные строки
        for column in df.columns:
            if column == 'Дата':  # Пропускаем столбец с датами
                continue
                
            # Определяем статью и контрагента
            if column.startswith('ДЗ:'):
                статья = 'Дебиторская задолженность'
                контрагент = column[3:].strip()  # Убираем префикс "ДЗ: "
            elif column.startswith('КЗ:'):
                статья = 'Кредиторская задолженность'
                контрагент = column[3:].strip()  # Убираем префикс "КЗ: "
            elif column == 'Дебиторская задолженность ИТОГО':
                статья = 'Дебиторская задолженность'
                контрагент = 'ИТОГО'
            elif column == 'Кредиторская задолженность ИТОГО':
                статья = 'Кредиторская задолженность'
                контрагент = 'ИТОГО'
            else:
                статья = column
                контрагент = 'Н/Д'  # Не применимо
            
            # Создаем временный DataFrame для текущего столбца
            temp_df = pd.DataFrame({
                'Дата': df['Дата'],
                'Статья': статья,
                'Контрагент': контрагент,
                'Сумма': df[column]
            })
            
            # Добавляем в общий DataFrame
            df_long = pd.concat([df_long, temp_df], ignore_index=True)
        
        # Сортируем по дате
        df_long = df_long.sort_values('Дата')
        
        # Записываем преобразованный DataFrame во вторую закладку
        df_long.to_excel(writer, sheet_name='Данные ЧОК (4 столбца)', index=False)
        
        # Форматирование второй закладки
        worksheet2 = writer.sheets['Данные ЧОК (4 столбца)']
        
        # Применяем форматирование к заголовкам
        for col_idx in range(1, 5):
            cell = worksheet2.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Автоподбор ширины столбцов
        worksheet2.column_dimensions['A'].width = 15  # Дата
        worksheet2.column_dimensions['B'].width = 25  # Статья
        worksheet2.column_dimensions['C'].width = 25  # Контрагент
        worksheet2.column_dimensions['D'].width = 15  # Сумма
        
        # Форматируем столбец с датами на втором листе
        for row_idx in range(2, len(df_long) + 2):
            cell = worksheet2.cell(row=row_idx, column=1)  # Столбец Дата (A)
            cell.number_format = 'mm.yyyy'
        
        # Форматирование числовых значений
        for row_idx in range(2, len(df_long) + 2):
            cell = worksheet2.cell(row=row_idx, column=4)  # Столбец Сумма
            cell.number_format = '#,##0.00₽'
    
    print(f"Данные успешно экспортированы в файл: {file_path}")

if __name__ == "__main__":
    # Создаем директорию, если она не существует
    os.makedirs(r"C:\OrangeDM_book_TS", exist_ok=True)
    
    # Генерируем набор данных
    print("Генерация набора месячных данных...")
    df = generate_dataset()
    
    # Проверяем наличие расчетных столбцов
    print("\nПроверка наличия расчетных столбцов:")
    print(f"Столбцы в датафрейме: {df.columns.tolist()}")
    print(f"Количество месяцев: {len(df)}")
    print(f"Первые 5 значений 'Запасы': {df['Запасы'].head().tolist()}")
    print(f"Первые 5 значений 'ЧОК': {df['ЧОК'].head().tolist()}")
    
    # Добавляем проверку диапазона дат
    print(f"\nДиапазон дат: от {df['Дата'].min().strftime('%m.%Y')} до {df['Дата'].max().strftime('%m.%Y')}")
    
    # Экспортируем в Excel
    print("\nЭкспорт данных в Excel...")
    export_to_excel(df)
    
    print("Готово!")

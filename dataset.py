import pandas as pd
import numpy as np
import datetime as dt
from dateutil.relativedelta import relativedelta
import random
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

def is_holiday_or_weekend(date):
    """Проверяет, является ли дата выходным или праздником в РФ"""
    # Выходные дни (суббота и воскресенье)
    if date.weekday() >= 5:
        return True
    
    # Основные российские праздники
    holidays = [
        # Новогодние каникулы
        (1, 1), (2, 1), (3, 1), (4, 1), (5, 1), (6, 1), (7, 1), (8, 1),
        # Рождество
        (7, 1),
        # День защитника отечества
        (23, 2),
        # Международный женский день
        (8, 3),
        # Праздник Весны и Труда
        (1, 5),
        # День Победы
        (9, 5),
        # День России
        (12, 6),
        # День народного единства
        (4, 11)
    ]
    
    # Проверка на праздники
    if (date.day, date.month) in holidays:
        return True
    
    # Дополнительно: перенос праздников, если они выпадают на выходные
    # (упрощенная логика, не учитывает все правила переносов)
    for day, month in holidays:
        holiday_date = dt.date(date.year, month, day)
        if holiday_date.weekday() >= 5:  # выпадает на выходной
            next_workday = holiday_date + dt.timedelta(days=(7 - holiday_date.weekday() + 1))
            if date == next_workday:
                return True
    
    return False

def generate_working_days(start_date, end_date):
    """Генерирует список рабочих дней между указанными датами"""
    working_days = []
    current_date = start_date
    while current_date <= end_date:
        if not is_holiday_or_weekend(current_date):
            working_days.append(current_date)
        current_date += dt.timedelta(days=1)
    return working_days

def generate_time_series_with_trend_and_seasonality(dates, base_value, trend_factor, seasonal_amplitude, noise_level):
    """
    Генерирует временной ряд с трендом и сезонностью
    
    Параметры:
    - dates: список дат
    - base_value: начальное значение
    - trend_factor: коэффициент тренда (годовой рост)
    - seasonal_amplitude: амплитуда сезонных колебаний
    - noise_level: уровень случайного шума
    """
    values = []
    for i, date in enumerate(dates):
        # Добавляем тренд (линейный рост)
        years_passed = (date - dates[0]).days / 365.25
        trend = base_value * (1 + trend_factor) ** years_passed
        
        # Добавляем сезонность (годовая)
        day_of_year = date.timetuple().tm_yday
        seasonality = seasonal_amplitude * np.sin(2 * np.pi * day_of_year / 365.25)
        
        # Добавляем случайный шум
        noise = np.random.normal(0, noise_level * trend)
        
        # Итоговое значение
        value = max(0, trend + seasonality * trend + noise)
        values.append(round(value, 2))
    
    return values

def generate_dataset():
    # Определяем период - 10 лет
    start_date = dt.date(2014, 1, 1)
    end_date = dt.date(2023, 12, 31)
    
    # Генерируем список рабочих дней
    working_days = generate_working_days(start_date, end_date)
    
    # Создаем базовый DataFrame с датами
    df = pd.DataFrame({'Дата': working_days})
    
    # Генерируем имена контрагентов
    debtor_names = [
        "ООО 'ТехноПром'", "АО 'Меркурий'", "ООО 'СтройИнвест'", 
        "ЗАО 'ЭнергоСбыт'", "ООО 'АгроХолдинг'", "ИП Иванов А.А.", 
        "ООО 'МеталлГрупп'", "АО 'ТрансЛогистика'", "ООО 'МедТех'", 
        "ЗАО 'ИТ-Решения'"
    ]
    
    creditor_names = [
        "ООО 'ПоставкаПлюс'", "АО 'ТехноСервис'", "ООО 'СырьеТорг'", 
        "ПАО 'ЭнергоСеть'", "ООО 'ЛогистикаПро'", "АО 'СтальИмпорт'", 
        "ООО 'ХимПром'", "ЗАО 'СтройМатериалы'", "ООО 'ФинансГрупп'", 
        "АО 'ТехИмпорт'"
    ]
    
    # Генерируем данные для материалов
    materials_base = 5_000_000
    df['Материалы'] = generate_time_series_with_trend_and_seasonality(
        working_days, materials_base, 0.08, 0.15, 0.02
    )
    
    # Генерируем данные для готовой продукции
    finished_goods_base = 7_000_000
    df['Готовая продукция'] = generate_time_series_with_trend_and_seasonality(
        working_days, finished_goods_base, 0.1, 0.2, 0.03
    )
    
    # Генерируем данные для дебиторской задолженности по контрагентам
    accounts_receivable_base = 10_000_000
    debtor_weights = [random.uniform(0.05, 0.2) for _ in range(10)]
    debtor_weights = [w/sum(debtor_weights) for w in debtor_weights]
    
    total_receivables = generate_time_series_with_trend_and_seasonality(
        working_days, accounts_receivable_base, 0.12, 0.1, 0.04
    )
    
    for i, name in enumerate(debtor_names):
        column_name = f"ДЗ: {name}"
        df[column_name] = [round(value * debtor_weights[i] * (1 + 0.1 * np.sin(j/30)), 2) 
                           for j, value in enumerate(total_receivables)]
    
    # Генерируем данные для кредиторской задолженности по контрагентам
    accounts_payable_base = 8_000_000
    creditor_weights = [random.uniform(0.05, 0.2) for _ in range(10)]
    creditor_weights = [w/sum(creditor_weights) for w in creditor_weights]
    
    total_payables = generate_time_series_with_trend_and_seasonality(
        working_days, accounts_payable_base, 0.07, 0.12, 0.03
    )
    
    for i, name in enumerate(creditor_names):
        column_name = f"КЗ: {name}"
        df[column_name] = [round(value * creditor_weights[i] * (1 + 0.1 * np.sin(j/45)), 2) 
                           for j, value in enumerate(total_payables)]
    
    # Добавляем итоговые столбцы
    df['Дебиторская задолженность ИТОГО'] = df[[col for col in df.columns if col.startswith('ДЗ:')]].sum(axis=1)
    df['Кредиторская задолженность ИТОГО'] = df[[col for col in df.columns if col.startswith('КЗ:')]].sum(axis=1)
    
    # Форматируем столбец даты
    df['Дата'] = pd.to_datetime(df['Дата'])
    
    return df

def export_to_excel(df, filename="financial_dataset.xlsx"):
    """Экспортирует данные в Excel с форматированием"""
    file_path = os.path.join(r"C:\OrangeDM_book_TS", filename)
    
    # Создаем писателя Excel
    with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Данные ЧОК', index=False)
        
        # Получаем доступ к листу для форматирования
        workbook = writer.book
        worksheet = writer.sheets['Данные ЧОК']
        
        # Форматируем заголовки
        header_font = Font(bold=True)
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Применяем форматирование к заголовкам
        for col_idx, column in enumerate(df.columns):
            cell = worksheet.cell(row=1, column=col_idx+1)
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Автоподбор ширины столбцов
        for col_idx, column in enumerate(df.columns):
            column_width = max(len(str(column)), 15)
            worksheet.column_dimensions[get_column_letter(col_idx+1)].width = column_width
        
        # Форматирование числовых значений
        for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, max_row=len(df)+1, min_col=2, max_col=len(df.columns))):
            for cell in row:
                cell.number_format = '#,##0.00₽'
    
    print(f"Данные успешно экспортированы в файл: {file_path}")

if __name__ == "__main__":
    # Создаем директорию, если она не существует
    os.makedirs(r"C:\OrangeDM_book_TS", exist_ok=True)
    
    # Генерируем набор данных
    print("Генерация набора данных...")
    df = generate_dataset()
    
    # Экспортируем в Excel
    print("Экспорт данных в Excel...")
    export_to_excel(df)
    
    print("Готово!")
